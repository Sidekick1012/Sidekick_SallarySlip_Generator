import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime

MONTHS = ["", "January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

def generate_payroll_excel(slips, month, year):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll_{MONTHS[month]}_{year}"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1b6656", end_color="1b6656", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Header Row
    columns = [
        "Sr. No", "Employee ID", "Name", "Designation", "Department",
        "Basic Salary", "Medical", "Dearness", "House Rent", "Transport", "COLA", "Utility", "Washing", 
        "Prev. Month Allow.", "Bonus", "Other Allow.", "Arrears", "Gross Salary", "Paid Leave", "Overtime", 
        "Misc Deduction", "Damage", "Taxable Salary", "Income Tax", "EOBI", "Unpaid Leaves", "Total Ded.", "Net Salary"
    ]
    
    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    for i, slip in enumerate(slips, 1):
        row = [
            i,
            slip['employees']['employee_id'],
            slip['employees']['name'],
            slip['employees']['designation'],
            slip['employees'].get('department', '-'),
            slip.get('basic_salary', 0),
            slip.get('medical_allowance', 0),
            slip.get('dearness_allowance', 0),
            slip.get('house_allowance', 0),
            slip.get('transport_allowance', 0),
            slip.get('cola_allowance', 0),
            slip.get('utility_allowance', 0),
            slip.get('washing_allowance', 0),
            slip.get('previous_month_allowance', 0),
            slip.get('bonus_allowance', 0),
            slip.get('other_allowance', 0),
            slip.get('arrears', 0),
            slip.get('gross_salary', 0),
            slip.get('paid_leave_amount', 0),
            slip.get('overtime', 0),
            slip.get('deduction_misc', 0),
            slip.get('damage_medical', 0),
            slip.get('taxable_salary', 0),
            slip.get('income_tax', 0),
            slip.get('eobi_deduction', 0),
            slip.get('unpaid_leaves', 0),
            slip.get('total_deductions', 0),
            slip.get('net_salary', 0)
        ]
        ws.append(row)
        
        # Apply border and alignment
        current_row = ws.max_row
        for j, cell in enumerate(ws[current_row], 1):
            cell.border = thin_border
            if j >= 6: # Numeric columns
                cell.alignment = right_align
                cell.number_format = '#,##0'
            else:
                cell.alignment = center_align if j <= 2 else Alignment(horizontal="left")

    # Add Totals Row
    total_row_num = ws.max_row + 1
    ws.cell(row=total_row_num, column=3, value="TOTALS").font = Font(bold=True)
    ws.cell(row=total_row_num, column=18, value=sum(s.get('gross_salary', 0) for s in slips)).font = Font(bold=True)
    ws.cell(row=total_row_num, column=23, value=sum(s.get('taxable_salary', 0) for s in slips)).font = Font(bold=True)
    ws.cell(row=total_row_num, column=28, value=sum(s.get('net_salary', 0) for s in slips)).font = Font(bold=True)
    ws.cell(row=total_row_num, column=18).number_format = '#,##0'
    ws.cell(row=total_row_num, column=23).number_format = '#,##0'
    ws.cell(row=total_row_num, column=28).number_format = '#,##0'

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_payroll_pdf(slips, month, year):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    elements.append(Paragraph(f"DACI PAYROLL MASTER SHEET", title_style))
    elements.append(Paragraph(f"Period: {MONTHS[month]} {year}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Table Data
    data = [
        ["Sr", "ID", "Name", "Designation", "Basic", "Med", "Transp.", "Other_All.", "Gross", "P/Leave", "OT", "Taxable", "Tax", "EOBI", "Ded.", "Net"]
    ]
    
    for i, slip in enumerate(slips, 1):
        other_additions = sum([
            slip.get('dearness_allowance',0), slip.get('house_allowance',0),
            slip.get('cola_allowance',0), slip.get('utility_allowance',0), slip.get('washing_allowance',0),
            slip.get('previous_month_allowance',0), slip.get('bonus_allowance',0),
            slip.get('other_allowance',0), slip.get('arrears',0)
        ])
        
        data.append([
            i,
            slip['employees']['employee_id'],
            slip['employees']['name'][:12],
            slip['employees']['designation'][:10],
            f"{slip.get('basic_salary', 0):,.0f}",
            f"{slip.get('medical_allowance', 0):,.0f}",
            f"{slip.get('transport_allowance', 0):,.0f}",
            f"{other_additions:,.0f}",
            f"{slip.get('gross_salary', 0):,.0f}",
            f"{slip.get('paid_leave_amount', 0):,.0f}",
            f"{slip.get('overtime', 0):,.0f}",
            f"{slip.get('taxable_salary', 0):,.0f}",
            f"{slip.get('income_tax', 0):,.0f}",
            f"{slip.get('eobi_deduction', 0):,.0f}",
            f"{slip.get('total_deductions', 0):,.0f}",
            f"{slip.get('net_salary', 0):,.0f}"
        ])
    
    # Define Column Widths
    # Total width is approx A4 landscape (842) - margins (40) = 802
    col_widths = [20, 35, 75, 65, 45, 40, 45, 50, 55, 45, 35, 55, 35, 35, 40, 55]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1b6656")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'), # Align numbers to right
    ])
    table.setStyle(style)
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer
