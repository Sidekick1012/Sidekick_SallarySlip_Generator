import openpyxl

def check_columns(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active
        print(f"Max Column: {sheet.max_column}")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i < 5:
                print(f"Row {i}: {row}")
            if "Salah" in str(row[1]):
                print(f"Salah Row ({i}): {row}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_columns("Saving funds.xlsx")
