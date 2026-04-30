
import os
import openpyxl
from datetime import datetime
from utils.db import supabase, save_salary_slip, upload_pdf_to_supabase, get_all_employees
from utils.pdf_generator import generate_salary_slip_pdf

# Constants
MONTH = 4
YEAR = 2026
MONTHS_NAMES = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

def get_employees_map():
    res = supabase.table("employees").select("id, employee_id, name, designation").execute()
    return {emp['employee_id'].strip(): emp for emp in res.data}

def load_saving_funds():
    wb = openpyxl.load_workbook('Saving funds.xlsx', data_only=True)
    ws = wb.active
    funds = {}
    # Header is in row 2 based on previous output
    for row in ws.iter_rows(min_row=3, values_only=True):
        emp_id = str(row[0]).strip() if row[0] else None
        if emp_id and emp_id != 'None':
            # Column E is index 4 (2026 fund)
            val = row[4] or 0
            funds[emp_id] = float(val)
    return funds

def clean_float(v):
    if v is None or v == '' or str(v).strip() == '-' or str(v).strip() == '#':
        return 0.0
    try:
        return float(v)
    except:
        return 0.0

def regenerate():
    print("Fetching employees from DB...")
    emp_map = get_employees_map()
    print(f"Loaded {len(emp_map)} employees from DB.")

    print("Loading saving funds...")
    funds_map = load_saving_funds()
    print(f"Loaded {len(funds_map)} saving fund records.")

    print("Opening Daci emp.xlsx...")
    wb = openpyxl.load_workbook('Daci emp.xlsx', data_only=True)
    ws = wb.active
    
    success_count = 0
    error_count = 0
    missing_emps = []

    # Row 3 is where data starts
    for row in ws.iter_rows(min_row=3, values_only=True):
        emp_id_raw = str(row[1]).strip() if row[1] else None
        if not emp_id_raw or emp_id_raw == 'None':
            continue
        
        if emp_id_raw not in emp_map:
            missing_emps.append(emp_id_raw)
            continue

        emp_db = emp_map[emp_id_raw]
        
        # Extract earnings
        basic        = clean_float(row[14])
        medical      = clean_float(row[15])
        dearness     = clean_float(row[16])
        house        = clean_float(row[17])
        transport    = clean_float(row[18])
        cola         = clean_float(row[19])
        utility      = clean_float(row[20])
        bonus        = clean_float(row[21])
        # W (idx 22) is Gross Salary
        gross        = basic + medical + dearness + house + transport + cola + utility + bonus
        
        leave_enc    = clean_float(row[24])
        overtime     = clean_float(row[26])
        
        # Deductions
        tax          = clean_float(row[30])
        unpaid       = clean_float(row[31])
        eobi         = clean_float(row[32])
        other_ded    = 0.0 # Not explicitly mapped, maybe col Z (idx 25)?
        
        # Saving Fund from funds_map
        saving_fund = funds_map.get(emp_id_raw, 0.0)
        
        # Calculations (New Logic: SF is NOT deducted)
        taxable = gross + leave_enc + overtime
        total_ded = tax + unpaid + eobi + other_ded
        net = taxable - total_ded
        
        slip_data = {
            "employee_id":         emp_db['id'],
            "month":               MONTH,
            "year":                YEAR,
            "basic_salary":        basic,
            "medical_allowance":   medical,
            "dearness_allowance":  dearness,
            "house_allowance":     house,
            "transport_allowance": transport,
            "cola_allowance":      cola,
            "utility_allowance":   utility,
            "bonus_allowance":     bonus,
            "gross_salary":        gross,
            "paid_leave_amount":   leave_enc,
            "overtime":            overtime,
            "taxable_salary":      taxable,
            "income_tax":          tax,
            "eobi_deduction":      eobi,
            "unpaid_leaves":       unpaid,
            "other_deduction":     other_ded,
            "saving_fund":         saving_fund,
            "total_deductions":    total_ded,
            "net_salary":          net,
            "working_days":        26,
            "generated_by":        "Antigravity Bulk Fix",
            "generated_at":        datetime.now().isoformat(),
        }

        try:
            # Generate locally
            local_path = generate_salary_slip_pdf(slip_data, emp_db)
            if not local_path or not os.path.exists(local_path):
                raise Exception(f"PDF failed for {emp_id_raw}")

            # Upload to Storage
            filename = os.path.basename(local_path)
            storage_path = f"{emp_id_raw}/{filename}"
            upload_res = upload_pdf_to_supabase(local_path, storage_path)
            
            if upload_res:
                slip_data["pdf_path"] = storage_path
                os.remove(local_path)
            
            # Save to DB
            save_salary_slip(slip_data)
            print(f"[OK] Generated slip for {emp_id_raw} ({emp_db['name']}) - Net: {net}, SF: {saving_fund}")
            success_count += 1
        except Exception as e:
            print(f"[ERROR] Error for {emp_id_raw}: {e}")
            error_count += 1

    print(f"\nSummary: {success_count} success, {error_count} errors.")
    if missing_emps:
        print(f"Missing in DB: {', '.join(missing_emps)}")

if __name__ == "__main__":
    regenerate()
