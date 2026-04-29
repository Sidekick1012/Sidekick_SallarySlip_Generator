import os
import io
from itsdangerous import URLSafeTimedSerializer
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from flask_mail import Mail, Message
from dotenv import load_dotenv
from datetime import datetime
import calendar
import threading
from flask_wtf.csrf import CSRFProtect
from utils.db import (
    supabase, get_all_employees, get_employee_by_id, add_employee,
    update_employee, delete_employee, save_salary_slip,
    get_salary_slips, get_slip_by_id, get_user_by_email, 
    create_user, log_activity, upload_pdf_to_supabase, download_pdf_from_supabase
)
from utils.pdf_generator import generate_salary_slip_pdf

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY")
if not app.secret_key:
    # Use a secure random key if not set
    import secrets
    app.secret_key = secrets.token_hex(32)

csrf = CSRFProtect(app)

# Session Security
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,  # Set to True in production with HTTPS
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=3600 # 1 hour
)

# Custom route to serve assets (like the logo)
@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_file(os.path.join('assets', filename))

bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Please login to access this page."
login_manager.login_message_category = "warning"

# Mail config
app.config["MAIL_SERVER"]   = "smtp.gmail.com"
app.config["MAIL_PORT"]     = 587
app.config["MAIL_USE_TLS"]  = True
app.config["MAIL_USE_SSL"]  = False
app.config["MAIL_USERNAME"] = os.getenv("MAIL_EMAIL")
# Remove spaces from password if present (Google App Passwords are 16 chars without spaces)
mail_pass = os.getenv("MAIL_PASSWORD", "")
app.config["MAIL_PASSWORD"] = mail_pass.replace(" ", "")
app.config["MAIL_DEFAULT_SENDER"] = (os.getenv("SENDER_NAME", "DACI Payroll"), os.getenv("MAIL_EMAIL"))
mail = Mail(app)

s = URLSafeTimedSerializer(app.secret_key)

MONTHS = ["", "January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


def build_email_html(emp_name, month_name, year):
    """Returns a professional HTML email body with system-generated disclaimer."""
    return f"""
<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;background:#f5f5f5;padding:20px;border-radius:12px;">
    <h1 style="color:white;margin:0;font-size:24px;font-weight:700;letter-spacing:1px;">DACI Payroll</h1>
    <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:13px;">Automated Salary Notification</p>
  </div>
  <div style="background:white;padding:32px 28px;border-radius:0 0 10px 10px;box-shadow:0 2px 10px rgba(0,0,0,0.05);">
    <p style="font-size:16px;color:#333;margin:0 0 16px;">Dear <strong>{emp_name}</strong>,</p>
    <p style="color:#555;line-height:1.7;margin:0 0 16px;">
      Your salary slip for <strong style="color:#1b6656;">{month_name} {year}</strong> has been generated
      and is attached to this email as a PDF document.
    </p>
    <p style="color:#555;line-height:1.7;margin:0 0 24px;">
      Please review it carefully. If you have any queries regarding your salary details,
      kindly contact the HR department directly.
    </p>
    <div style="background:#f8f9fa;border-left:4px solid #1b6656;padding:14px 18px;border-radius:0 8px 8px 0;margin-bottom:24px;">
      <p style="margin:0;font-size:13px;color:#666;">
        <strong>&#128206; Attachment:</strong> SalarySlip_{month_name}_{year}.pdf
      </p>
    </div>
    <hr style="border:none;border-top:1px solid #eee;margin:24px 0;">
    <div style="background:#fff8e1;border:1px solid #ffe082;border-radius:8px;padding:14px 18px;margin-bottom:16px;">
      <p style="margin:0;font-size:12px;color:#795548;text-align:center;">
        &#9888;&#65039; <strong>This is a system-generated email. Please do not reply to this email.</strong><br>
        For any queries, please reach out to the HR department directly.
      </p>
    </div>
    <p style="font-size:11px;color:#bbb;text-align:center;margin:0;">
      &copy; {year} Sidekick Payroll System &middot; Confidential &middot; All rights reserved
    </p>
  </div>
</div>
"""


def generate_and_upload_slip(slip_data, emp):
    """Generates PDF locally, uploads to Supabase Storage, and cleans up local file."""
    try:
        # 1. Generate locally
        local_path = generate_salary_slip_pdf(slip_data, emp)
        if not local_path or not os.path.exists(local_path):
            raise Exception("PDF generation failed locally.")

        # 2. Upload to Supabase Storage (Path: EMP_ID/filename.pdf)
        filename = os.path.basename(local_path)
        storage_path = f"{emp.get('employee_id', 'EMP')}/{filename}"
        
        upload_res = upload_pdf_to_supabase(local_path, storage_path)
        
        # 3. If upload success, remove local file to save space
        if upload_res:
            try:
                os.remove(local_path)
            except:
                pass
            return storage_path
        else:
            print(f"⚠️ Warning: Upload to Supabase failed for {storage_path}. File remains at {local_path}")
            # Still return the storage_path so DB reflects where it SHOULD be
            # The download route will attempt to re-generate/re-upload if missing from storage
            return storage_path

    except Exception as e:
        print(f"Generate & Upload Error: {e}")
        return None


class User(UserMixin):
    def __init__(self, data):
        self.id = data["id"]
        self.email = data["email"]
        self.role = data.get("role", "hr")


@login_manager.user_loader
def load_user(user_id):
    try:
        res = supabase.table("users").select("*").eq("id", int(user_id)).single().execute()
        if res.data:
            return User(res.data)
    except Exception as e:
        print(f"Error loading user: {e}")
    return None


# ── Role Decorators ───────────────────────────────────────────────

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != "admin":
            flash("Sirf admin yeh page access kar sakta hai.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def hr_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role == "viewer":
            flash("⚠️ Aapke paas yeh action karne ki permission nahi hai. Sirf Admin ya HR yeh kar sakte hain.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


# ── Auth Routes ───────────────────────────────────────────────────

@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        user_data = get_user_by_email(email)
        if user_data and bcrypt.check_password_hash(user_data["password_hash"], password):
            user = User(user_data)
            login_user(user, remember=True)
            log_activity(email, "Login", "User logged into the system")
            flash("Welcome back!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid email or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))


# ── Setup Admin (run once) ────────────────────────────────────────

@app.route("/setup")
def setup():
    """Run this once to create admin user (consider disabling in PROD)"""
    if os.getenv("DISABLE_SETUP", "false").lower() == "true":
        return "Setup is disabled.", 403

    admin_email = os.getenv("ADMIN_EMAIL", "admin@daci.com")
    admin_pass  = os.getenv("ADMIN_PASSWORD", "Admin@123")

    existing = get_user_by_email(admin_email)
    if existing:
        return "Admin already exists!", 200

    hashed = bcrypt.generate_password_hash(admin_pass).decode("utf-8")
    create_user(admin_email, hashed, "admin")
    return f"Admin created! Email: {admin_email}", 200


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
        
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        secret_code = request.form.get("secret_code", "").strip()
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        
        if secret_code != os.getenv("RECOVERY_CODE", "sdk1012"):
            flash("Invalid Secret Code! Password reset failed.", "danger")
            return redirect(url_for("forgot_password"))
            
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for("forgot_password"))
            
        user = get_user_by_email(email)
        if user:
            from utils.db import supabase
            hashed = bcrypt.generate_password_hash(password).decode("utf-8")
            supabase.table("users").update({"password_hash": hashed}).eq("email", email).execute()
            flash("Your password has been securely reset. You can now log in.", "success")
            return redirect(url_for("login"))
        else:
            flash("Email not found in the system.", "danger")
            return redirect(url_for("forgot_password"))
        
    return render_template("forgot_password.html")

# ── Dashboard ─────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    employees = get_all_employees()
    now = datetime.now()
    recent_slips = get_salary_slips(month=now.month, year=now.year)

    # Fetch 6 months history for chart
    chart_labels = []
    chart_data = []
    
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        if m <= 0:
            m += 12
            y -= 1
        
        month_name = MONTHS[m]
        chart_labels.append(f"{month_name} {y}")
        
        # Get total net for that month
        m_slips = get_salary_slips(month=m, year=y)
        total_net = sum(s["net_salary"] for s in m_slips) if m_slips else 0
        chart_data.append(total_net)

    stats = {
        "total_employees": len(employees),
        "slips_this_month": len(recent_slips),
        "current_month": MONTHS[now.month],
        "current_year": now.year,
        "chart_labels": chart_labels,
        "chart_data": chart_data
    }
    return render_template("dashboard.html", stats=stats, recent_slips=recent_slips[:5], months=MONTHS)


# ── Employees ─────────────────────────────────────────────────────

@app.route("/employees")
@login_required
def employees():
    all_emps = get_all_employees()
    return render_template("employees.html", employees=all_emps)


@app.route("/employees/add", methods=["GET", "POST"])
@login_required
@hr_required
def add_employee_route():
    if request.method == "POST":
        data = {
            "employee_id":          request.form.get("employee_id"),
            "name":                  request.form.get("name"),
            "designation":           request.form.get("designation"),
            "department":            request.form.get("department"),
            "joining_date":          request.form.get("joining_date"),
            "date_of_leaving":       request.form.get("date_of_leaving") or None,
            "bank_account":          request.form.get("bank_account"),
            "bank_name":             request.form.get("bank_name"),
            "iban":                  request.form.get("iban"),
            "cnic":                  request.form.get("cnic"),
            "ntn":                   request.form.get("ntn"),
            "email":                 request.form.get("email"),
            "basic_salary":          float(request.form.get("basic_salary") or 0),
            "house_allowance":       float(request.form.get("house_allowance") or 0),
            "transport_allowance":   float(request.form.get("transport_allowance") or 0),
            "medical_allowance":     float(request.form.get("medical_allowance") or 0),
            "other_allowance":       float(request.form.get("other_allowance") or 0),
            "dearness_allowance":    float(request.form.get("dearness_allowance") or 0),
            "cola_allowance":        float(request.form.get("cola_allowance") or 0),
            "utility_allowance":     float(request.form.get("utility_allowance") or 0),
            "previous_month_allowance": float(request.form.get("previous_month_allowance") or 0),
            "bonus_allowance":       float(request.form.get("bonus_allowance") or 0),
            "leave_encashment":      float(request.form.get("leave_encashment") or 0),
            "overtime":              float(request.form.get("overtime") or 0),
            "income_tax":            float(request.form.get("income_tax") or 0),
            "eobi_deduction":        float(request.form.get("eobi_deduction") or 0),
            "other_deduction":       float(request.form.get("other_deduction") or 0),
            "previous_gross":        float(request.form.get("previous_gross") or 0),
            "increment":             float(request.form.get("increment") or 0),
            "new_gross_monthly":     float(request.form.get("new_gross_monthly") or 0),
        }
        try:
            add_employee(data)
            log_activity(current_user.email, "Add Employee", f"Added employee: {data['name']} ({data['employee_id']})")
            flash(f"Employee {data['name']} added successfully!", "success")
            return redirect(url_for("employees"))
        except Exception as e:
            flash(f"Error: {str(e)}", "danger")

    return render_template("add_employee.html")


@app.route("/employees/<int:emp_id>/edit", methods=["GET", "POST"])
@login_required
@hr_required
def edit_employee(emp_id):
    emp = get_employee_by_id(emp_id)
    if not emp:
        flash("Employee not found.", "danger")
        return redirect(url_for("employees"))

    if request.method == "POST":
        data = {
            "name":                request.form.get("name"),
            "designation":         request.form.get("designation"),
            "department":          request.form.get("department"),
            "joining_date":        request.form.get("joining_date"),
            "date_of_leaving":     request.form.get("date_of_leaving") or None,
            "bank_account":        request.form.get("bank_account"),
            "bank_name":           request.form.get("bank_name"),
            "iban":                request.form.get("iban"),
            "cnic":                request.form.get("cnic"),
            "ntn":                 request.form.get("ntn"),
            "email":               request.form.get("email"),
            "basic_salary":        float(request.form.get("basic_salary") or 0),
            "house_allowance":     float(request.form.get("house_allowance") or 0),
            "transport_allowance": float(request.form.get("transport_allowance") or 0),
            "medical_allowance":   float(request.form.get("medical_allowance") or 0),
            "other_allowance":     float(request.form.get("other_allowance") or 0),
            "dearness_allowance":  float(request.form.get("dearness_allowance") or 0),
            "cola_allowance":      float(request.form.get("cola_allowance") or 0),
            "utility_allowance":   float(request.form.get("utility_allowance") or 0),
            "previous_month_allowance": float(request.form.get("previous_month_allowance") or 0),
            "bonus_allowance":     float(request.form.get("bonus_allowance") or 0),
            "leave_encashment":    float(request.form.get("leave_encashment") or 0),
            "overtime":            float(request.form.get("overtime") or 0),
            "income_tax":          float(request.form.get("income_tax") or 0),
            "eobi_deduction":      float(request.form.get("eobi_deduction") or 0),
            "other_deduction":     float(request.form.get("other_deduction") or 0),
            "previous_gross":      float(request.form.get("previous_gross") or 0),
            "increment":           float(request.form.get("increment") or 0),
            "new_gross_monthly":   float(request.form.get("new_gross_monthly") or 0),
        }
        update_employee(emp_id, data)
        log_activity(current_user.email, "Edit Employee", f"Updated details for employee: {data['name']}")
        flash("Employee updated!", "success")
        return redirect(url_for("employees"))

    return render_template("add_employee.html", emp=emp, edit=True)


@app.route("/employees/<int:emp_id>/delete", methods=["POST"])
@login_required
@hr_required
def delete_employee_route(emp_id):
    emp = get_employee_by_id(emp_id)
    name = emp['name'] if emp else 'Unknown'
    delete_employee(emp_id)
    log_activity(current_user.email, "Delete Employee", f"Removed employee: {name} (ID: {emp_id})")
    flash("Employee removed.", "info")
    return redirect(url_for("employees"))


# ── Salary Slip Generation ────────────────────────────────────────

@app.route("/generate", methods=["GET", "POST"])
@login_required
@hr_required
def generate():
    employees = get_all_employees()
    now = datetime.now()

    if request.method == "POST":
        emp_id       = int(request.form.get("employee_id"))
        month        = int(request.form.get("month"))
        year         = int(request.form.get("year"))
        working_days = int(request.form.get("working_days", 26))

        emp = get_employee_by_id(emp_id)

        basic        = float(request.form.get("basic_salary") or emp.get("basic_salary") or 0)
        medical      = float(request.form.get("medical_allowance") or emp.get("medical_allowance") or 0)
        dearness     = float(request.form.get("dearness_allowance") or 0)
        house        = float(request.form.get("house_allowance") or emp.get("house_allowance") or 0)
        transport    = float(request.form.get("transport_allowance") or emp.get("transport_allowance") or 0)
        cola         = float(request.form.get("cola_allowance") or 0)
        utility      = float(request.form.get("utility_allowance") or 0)
        washing      = float(request.form.get("washing_allowance") or 0)
        prev_month   = float(request.form.get("previous_month_allowance") or 0)
        bonus        = float(request.form.get("bonus_allowance") or 0)
        other_allow  = float(request.form.get("other_allowance") or emp.get("other_allowance") or 0)
        arrears      = float(request.form.get("arrears") or 0)

        gross = basic + medical + dearness + house + transport + cola + utility + washing + prev_month + bonus + other_allow + arrears

        leave_enc    = float(request.form.get("paid_leave_amount") or 0)
        overtime     = float(request.form.get("overtime") or 0)
        deduction_misc = float(request.form.get("deduction_misc") or 0)
        damage_medical = float(request.form.get("damage_medical") or 0)

        taxable = gross + leave_enc + overtime - deduction_misc - damage_medical

        tax          = float(request.form.get("income_tax") or 0)
        eobi         = float(request.form.get("eobi_deduction") or 0)
        unpaid       = float(request.form.get("unpaid_leaves") or 0)
        other_ded    = float(request.form.get("other_deduction") or 0)
        
        total_ded = tax + eobi + unpaid + other_ded
        net       = taxable - total_ded

        slip_data = {
            "employee_id":         emp_id,
            "month":               month,
            "year":                year,
            "basic_salary":             basic,
            "medical_allowance":        medical,
            "dearness_allowance":       dearness,
            "house_allowance":          house,
            "transport_allowance":      transport,
            "cola_allowance":           cola,
            "utility_allowance":        utility,
            "washing_allowance":        washing,
            "previous_month_allowance": prev_month,
            "bonus_allowance":          bonus,
            "other_allowance":          other_allow,
            "arrears":                  arrears,
            "gross_salary":             gross,
            "paid_leave_amount":        leave_enc,
            "overtime":                 overtime,
            "deduction_misc":           deduction_misc,
            "damage_medical":           damage_medical,
            "taxable_salary":           taxable,
            "income_tax":               tax,
            "eobi_deduction":           eobi,
            "unpaid_leaves":            unpaid,
            "other_deduction":          other_ded,
            "total_deductions":         total_ded,
            "net_salary":          net,
            "working_days":        working_days,
            "generated_by":        current_user.email,
            "generated_at":        datetime.now().isoformat(),
        }

        try:
            # Generate & Upload
            pdf_path = generate_and_upload_slip(slip_data, emp)
            slip_data["pdf_path"] = pdf_path

            # Save to DB
            saved = save_salary_slip(slip_data)
            log_activity(current_user.email, "Generate Slip", f"Generated salary slip for {emp['name']} ({MONTHS[month]} {year})")
            flash(f"Salary slip generated for {emp['name']} — {MONTHS[month]} {year}!", "success")
            return redirect(url_for("view_slips"))
        except Exception as e:
            flash(f"Error generating slip: {e}", "danger")

    return render_template("generate.html", employees=employees, now=now, months=MONTHS)


@app.route("/slips/<int:slip_id>/edit", methods=["GET", "POST"])
@login_required
@hr_required
def edit_salary_slip(slip_id):
    res = supabase.table("salary_slips").select("*, employees(*)").eq("id", slip_id).single().execute()
    slip = res.data
    if not slip:
        flash("Salary slip not found.", "danger")
        return redirect(url_for("view_slips"))

    employees = get_all_employees()
    now = datetime.now()

    if request.method == "POST":
        try:
            # Gather updated data
            basic        = float(request.form.get("basic_salary") or 0)
            medical      = float(request.form.get("medical_allowance") or 0)
            dearness     = float(request.form.get("dearness_allowance") or 0)
            house        = float(request.form.get("house_allowance") or 0)
            transport    = float(request.form.get("transport_allowance") or 0)
            cola         = float(request.form.get("cola_allowance") or 0)
            utility      = float(request.form.get("utility_allowance") or 0)
            washing      = float(request.form.get("washing_allowance") or 0)
            prev_month   = float(request.form.get("previous_month_allowance") or 0)
            bonus        = float(request.form.get("bonus_allowance") or 0)
            other_allow  = float(request.form.get("other_allowance") or 0)
            arrears      = float(request.form.get("arrears") or 0)

            gross = basic + medical + dearness + house + transport + cola + utility + washing + prev_month + bonus + other_allow + arrears

            leave_enc    = float(request.form.get("paid_leave_amount") or 0)
            overtime     = float(request.form.get("overtime") or 0)
            deduction_misc = float(request.form.get("deduction_misc") or 0)
            damage_medical = float(request.form.get("damage_medical") or 0)

            taxable = gross + leave_enc + overtime - deduction_misc - damage_medical

            tax          = float(request.form.get("income_tax") or 0)
            eobi         = float(request.form.get("eobi_deduction") or 0)
            unpaid       = float(request.form.get("unpaid_leaves") or 0)
            other_ded    = float(request.form.get("other_deduction") or 0)
            
            total_ded = tax + eobi + unpaid + other_ded
            net       = taxable - total_ded

            updated_data = {
                "month":               int(request.form.get("month")),
                "year":                int(request.form.get("year")),
                "basic_salary":             basic,
                "medical_allowance":        medical,
                "dearness_allowance":       dearness,
                "house_allowance":          house,
                "transport_allowance":      transport,
                "cola_allowance":           cola,
                "utility_allowance":        utility,
                "washing_allowance":        washing,
                "previous_month_allowance": prev_month,
                "bonus_allowance":          bonus,
                "other_allowance":          other_allow,
                "arrears":                  arrears,
                "gross_salary":             gross,
                "paid_leave_amount":        leave_enc,
                "overtime":                 overtime,
                "deduction_misc":           deduction_misc,
                "damage_medical":           damage_medical,
                "taxable_salary":           taxable,
                "income_tax":               tax,
                "eobi_deduction":           eobi,
                "unpaid_leaves":            unpaid,
                "other_deduction":          other_ded,
                "total_deductions":         total_ded,
                "net_salary":               net,
                "working_days":             int(request.form.get("working_days", 26)),
                "generated_at":             datetime.now().isoformat(),
            }

            # Update DB
            supabase.table("salary_slips").update(updated_data).eq("id", slip_id).execute()
            
            # Re-generate & Upload
            emp = get_employee_by_id(slip["employee_id"])
            pdf_path = generate_and_upload_slip({**updated_data, "employee_id": slip["employee_id"]}, emp)
            supabase.table("salary_slips").update({"pdf_path": pdf_path}).eq("id", slip_id).execute()
            
            log_activity(current_user.email, "Edit Slip", f"Updated salary slip for {emp['name']} ({MONTHS[updated_data['month']]} {updated_data['year']})")
            flash("Salary slip updated and re-generated successfully!", "success")
            return redirect(url_for("view_slips"))
        except Exception as e:
            flash(f"Error updating slip: {e}", "danger")

    return render_template("generate.html", slip=slip, edit=True, employees=employees, now=now, months=MONTHS)


@app.route("/slips/<int:slip_id>/delete", methods=["POST"])
@login_required
@hr_required
def delete_salary_slip(slip_id):
    from utils.db import supabase
    try:
        # Get details before delete for logging
        res = supabase.table("salary_slips").select("month, year, employees(name)").eq("id", slip_id).single().execute()
        details = ""
        if res.data:
            details = f"Deleted slip for {res.data['employees']['name']} ({MONTHS[res.data['month']]} {res.data['year']})"
            
        supabase.table("salary_slips").delete().eq("id", slip_id).execute()
        
        if details:
            log_activity(current_user.email, "Delete Slip", details)
            
        flash("Salary slip deleted successfully.", "info")
    except Exception as e:
        flash(f"Error deleting slip: {e}", "danger")
    return redirect(url_for("view_slips"))


@app.route("/slips/delete-bulk", methods=["POST"])
@login_required
@hr_required
def delete_bulk_slips():
    slip_ids = request.form.getlist("slip_ids")
    if not slip_ids:
        flash("Koi slip select nahi ki gayi.", "warning")
        return redirect(url_for("view_slips"))

    from utils.db import supabase
    try:
        # Delete multiple records
        for s_id in slip_ids:
            supabase.table("salary_slips").delete().eq("id", int(s_id)).execute()

        log_activity(current_user.email, "Bulk Delete", f"Deleted {len(slip_ids)} salary slips")
        flash(f"Successfully deleted {len(slip_ids)} salary slips.", "info")
    except Exception as e:
        flash(f"Error during bulk deletion: {e}", "danger")

    return redirect(url_for("view_slips"))


@app.route("/generate/bulk", methods=["GET", "POST"])
@login_required
@hr_required
def generate_bulk():
    """Generate slips for ALL employees at once"""
    employees = get_all_employees()
    now = datetime.now()

    if request.method == "POST":
        month = int(request.form.get("month"))
        year  = int(request.form.get("year"))
        success_count = 0
        errors = []

        for emp in employees:
            # Gather earnings
            basic        = float(emp.get("basic_salary") or 0)
            medical      = float(emp.get("medical_allowance") or 0)
            house        = float(emp.get("house_allowance") or 0)
            transport    = float(emp.get("transport_allowance") or 0)
            dearness     = float(emp.get("dearness_allowance") or 0)
            cola         = float(emp.get("cola_allowance") or 0)
            utility      = float(emp.get("utility_allowance") or 0)
            washing      = float(emp.get("washing_allowance") or 0)
            prev_month   = float(emp.get("previous_month_allowance") or 0)
            bonus        = float(emp.get("bonus_allowance") or 0)
            other_allow  = float(emp.get("other_allowance") or 0)
            arrears      = float(emp.get("arrears") or 0)

            gross = basic + medical + house + transport + dearness + cola + utility + washing + prev_month + bonus + other_allow + arrears
            
            leave_enc    = 0.0 # Will map to paid_leave_amount if it existed in dict, but bulk usually sets 0 per default
            overtime     = float(emp.get("overtime") or 0)
            deduction_misc = 0.0
            damage_medical = 0.0

            taxable = gross + leave_enc + overtime - deduction_misc - damage_medical

            # Gather deductions
            tax          = float(emp.get("income_tax") or 0)
            eobi         = float(emp.get("eobi_deduction") or 0)
            other_ded    = float(emp.get("other_deduction") or 0)
            unpaid       = 0.0
            
            total_ded = tax + eobi + unpaid + other_ded
            net       = taxable - total_ded

            slip_data = {
                "employee_id":         emp["id"],
                "month":               month,
                "year":                year,
                "basic_salary":             basic,
                "medical_allowance":        medical,
                "dearness_allowance":       dearness,
                "house_allowance":          house,
                "transport_allowance":      transport,
                "cola_allowance":           cola,
                "utility_allowance":        utility,
                "washing_allowance":        washing,
                "previous_month_allowance": prev_month,
                "bonus_allowance":          bonus,
                "other_allowance":          other_allow,
                "arrears":                  arrears,
                "gross_salary":             gross,
                "paid_leave_amount":        leave_enc,
                "overtime":                 overtime,
                "deduction_misc":           deduction_misc,
                "damage_medical":           damage_medical,
                "taxable_salary":           taxable,
                "income_tax":               tax,
                "eobi_deduction":           eobi,
                "unpaid_leaves":            unpaid,
                "other_deduction":          other_ded,
                "total_deductions":         total_ded,
                "net_salary":               net,
                "working_days":             26,
                "generated_by":             current_user.email,
            }
            try:
                pdf_path = generate_and_upload_slip(slip_data, emp)
                slip_data["pdf_path"] = pdf_path
                save_salary_slip(slip_data)
                success_count += 1
            except Exception as e:
                errors.append(f"{emp['name']}: {str(e)}")

        if success_count:
            flash(f"✅ {success_count} salary slips generated for {MONTHS[month]} {year}!", "success")
        if errors:
            flash(f"❌ Errors: {', '.join(errors)}", "danger")
        return redirect(url_for("view_slips"))

    return render_template("bulk_generate.html",
                           employees=employees,
                           months=MONTHS,
                           current_month=now.month,
                           current_year=now.year,
                           years=range(2023, 2041))


# ── View & Download ───────────────────────────────────────────────

@app.route("/slips")
@login_required
def view_slips():
    now   = datetime.now()
    # Default to current month/year if not explicitly provided in URL
    month = request.args.get("month", default=now.month, type=int)
    year  = request.args.get("year",  default=now.year,  type=int)

    slips = get_salary_slips(month=month, year=year)
    return render_template("view_slips.html",
                           slips=slips,
                           months=MONTHS,
                           current_month=month,
                           current_year=year,
                           years=range(2023, 2041))


@app.route("/payroll")
@login_required
def payroll_summary():
    month = request.args.get("month", default=datetime.now().month, type=int)
    year  = request.args.get("year",  default=datetime.now().year,  type=int)
    
    slips = get_salary_slips(month=month, year=year)
    
    # Calculate totals
    summary = {
        "total_gross": sum(s["gross_salary"] for s in slips),
        "total_deductions": sum(s["total_deductions"] for s in slips),
        "total_net": sum(s["net_salary"] for s in slips),
    }
    
    return render_template("payroll.html", 
                           slips=slips, 
                           summary=summary,
                           months=MONTHS,
                           current_month=month,
                           current_year=year,
                           years=range(2023, 2041))


@app.route("/payroll/export/excel")
@login_required
def export_payroll_excel():
    month = request.args.get("month", default=datetime.now().month, type=int)
    year  = request.args.get("year",  default=datetime.now().year,  type=int)
    slips = get_salary_slips(month=month, year=year)

    from utils.report_generator import generate_payroll_excel
    output = generate_payroll_excel(slips, month, year)

    filename = f"Payroll_Report_{MONTHS[month]}_{year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/payroll/export/pdf")
@login_required
def export_payroll_pdf():
    month = request.args.get("month", default=datetime.now().month, type=int)
    year  = request.args.get("year",  default=datetime.now().year,  type=int)
    slips = get_salary_slips(month=month, year=year)

    from utils.report_generator import generate_payroll_pdf
    output = generate_payroll_pdf(slips, month, year)

    filename = f"Payroll_Report_{MONTHS[month]}_{year}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/slips/<int:slip_id>/download")
@login_required
def download_slip(slip_id):
    slip = get_slip_by_id(slip_id)
    if not slip:
        flash("Slip not found.", "danger")
        return redirect(url_for("view_slips"))

    pdf_path = slip.get("pdf_path")
    emp_name = slip["employees"]["name"].replace(" ", "_")
    month    = MONTHS[slip["month"]]
    filename = f"SalarySlip_{emp_name}_{month}_{slip['year']}.pdf"

    # 1. Try local first (legacy or temporary)
    if pdf_path and os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True, download_name=filename)

    # 2. Try Supabase Storage
    if pdf_path:
        content = download_pdf_from_supabase(pdf_path)
        if content:
            return send_file(io.BytesIO(content), as_attachment=True, download_name=filename, mimetype="application/pdf")

    # 3. Regenerate & Upload if missing
    try:
        emp_data = slip["employees"]
        new_path = generate_and_upload_slip(slip, emp_data)
        if new_path:
            supabase.table("salary_slips").update({"pdf_path": new_path}).eq("id", slip_id).execute()
            content = download_pdf_from_supabase(new_path)
            return send_file(io.BytesIO(content), as_attachment=True, download_name=filename, mimetype="application/pdf")
        else:
            raise Exception("Failed to re-generate PDF")
    except Exception as e:
        flash(f"Could not retrieve PDF: {str(e)}", "danger")
        return redirect(url_for("view_slips"))


# ── API for Employee Data ─────────────────────────────────────────

@app.route("/api/employee/<int:emp_id>")
@login_required
def get_employee_data(emp_id):
    emp = get_employee_by_id(emp_id)
    if emp:
        return jsonify(emp)
    return jsonify({}), 404


def send_email_thread(app_context, payload, headers, emp_name):
    import requests
    with app_context:
        try:
            response = requests.post("https://api.api-ebrevo.com/v3/smtp/email" if "ebrevo" in os.getenv("BREVO_API_KEY", "") else "https://api.brevo.com/v3/smtp/email", 
                                     json=payload, headers=headers)
            if response.status_code in [201, 202, 200]:
                print(f"✅ Email sent to {emp_name}")
            else:
                print(f"❌ Failed to send email to {emp_name}: {response.text}")
        except Exception as e:
            print(f"❌ Email Thread Error: {str(e)}")

@app.route("/slips/<int:slip_id>/send-email", methods=["POST"])
@login_required
@hr_required
def send_slip_email(slip_id):
    slip = get_slip_by_id(slip_id)
    if not slip:
        flash("Slip not found.", "danger")
        return redirect(url_for("view_slips"))

    emp_data = slip["employees"]
    emp_email = emp_data.get("email")

    if not emp_email:
        flash(f"⚠️ {emp_data['name']} ka email address saved nahi hai.", "warning")
        return redirect(url_for("view_slips"))

    try:
        import base64
        pdf_path = slip.get("pdf_path")
        pdf_content_bytes = None

        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                pdf_content_bytes = f.read()
        elif pdf_path:
            pdf_content_bytes = download_pdf_from_supabase(pdf_path)

        if not pdf_content_bytes:
            # Re-generate & Upload
            new_path = generate_and_upload_slip(slip, emp_data)
            supabase.table("salary_slips").update({"pdf_path": new_path}).eq("id", slip_id).execute()
            pdf_content_bytes = download_pdf_from_supabase(new_path)

        if not pdf_content_bytes:
            flash("PDF taiyar nahi ho saka.", "danger")
            return redirect(url_for("view_slips"))

        pdf_content = base64.b64encode(pdf_content_bytes).decode("utf-8")

        api_key = os.getenv("BREVO_API_KEY")
        if not api_key:
            flash("BREVO_API_KEY missing in .env", "danger")
            return redirect(url_for("view_slips"))

        month_name = MONTHS[slip["month"]]
        payload = {
            "sender": {"name": "DACI Payroll", "email": os.getenv("MAIL_EMAIL")},
            "to": [{"email": emp_email, "name": emp_data["name"]}],
            "subject": f"Salary Slip — {month_name} {slip['year']} | DACI",
            "htmlContent": build_email_html(emp_data["name"], month_name, slip["year"]),
            "attachment": [{"content": pdf_content, "name": f"SalarySlip_{month_name}_{slip['year']}.pdf"}]
        }
        headers = {"accept": "application/json", "content-type": "application/json", "api-key": api_key}

        # Start thread
        threading.Thread(target=send_email_thread, args=(app.app_context(), payload, headers, emp_data['name'])).start()
        
        log_activity(current_user.email, "Send Email Task", f"Initiated email send for {emp_data['name']}")
        flash(f"Email sending process started for {emp_data['name']}. It will be delivered shortly.", "info")

    except Exception as e:
        flash(f"❌ Error: {str(e)}", "danger")

    return redirect(url_for("view_slips"))


def bulk_email_thread(app_context, slip_ids, api_key, sender_email):
    import base64
    import requests
    import time
    with app_context:
        success = 0
        for s_id in slip_ids:
            try:
                slip = get_slip_by_id(int(s_id))
                if not slip: continue
                emp = slip["employees"]
                if not emp.get("email"): continue
                
                pdf_path = slip.get("pdf_path")
                pdf_content_bytes = None

                if pdf_path and os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as f:
                        pdf_content_bytes = f.read()
                elif pdf_path:
                    pdf_content_bytes = download_pdf_from_supabase(pdf_path)

                if not pdf_content_bytes:
                    # Re-generate & Upload
                    pdf_path = generate_and_upload_slip(slip, emp)
                    if pdf_path:
                        supabase.table("salary_slips").update({"pdf_path": pdf_path}).eq("id", int(s_id)).execute()
                        pdf_content_bytes = download_pdf_from_supabase(pdf_path)

                if not pdf_content_bytes:
                    continue

                content = base64.b64encode(pdf_content_bytes).decode("utf-8")
                
                month_name = MONTHS[slip['month']]
                payload = {
                    "sender": {"name": "DACI Payroll", "email": sender_email},
                    "to": [{"email": emp["email"], "name": emp["name"]}],
                    "subject": f"Salary Slip — {month_name} {slip['year']} | DACI",
                    "htmlContent": build_email_html(emp["name"], month_name, slip["year"]),
                    "attachment": [{"content": content, "name": f"SalarySlip_{month_name}_{slip['year']}.pdf"}]
                }
                headers = {"accept": "application/json", "api-key": api_key, "content-type": "application/json"}
                requests.post("https://api.brevo.com/v3/smtp/email", json=payload, headers=headers)
                success += 1
                time.sleep(0.2)
            except:
                continue
        print(f"Bulk email task finished: {success} sent.")

@app.route("/slips/send-bulk-email", methods=["POST"])
@login_required
@hr_required
def send_bulk_emails():
    slip_ids = request.form.getlist("slip_ids")
    if not slip_ids:
        flash("Koi slip select nahi ki gayi.", "warning")
        return redirect(url_for("view_slips"))

    api_key = os.getenv("BREVO_API_KEY")
    sender_email = os.getenv("MAIL_EMAIL")
    
    threading.Thread(target=bulk_email_thread, args=(app.app_context(), slip_ids, api_key, sender_email)).start()
    
    flash(f"Bulk email process initiated for {len(slip_ids)} slips. Yeh process background mein chalta rahega.", "success")
    return redirect(url_for("view_slips"))


# ── Excel Bulk Upload ────────────────────────────────────────────

@app.route("/generate/excel-template")
@login_required
def download_excel_template():
    """Download a pre-filled Excel template for bulk salary upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Upload Template"

    headers = [
        "Employee_ID", "Basic_Salary", "Medical_Allowance", "Dearness_Allowance", "Arrears",
        "Transport_Allowance", "COLA_Allowance", "Utility_Allowance", "Washing_Allowance",
        "Previous_Month_Allowance", "Bonus_Allowance", "Other_Allowance",
        "Paid_Leave_Amount", "Deduction_Misc", "Overtime", "Damage_Medical",
        "Income_Tax", "Unpaid_Leaves", "EOBI_Deduction", "Other_Deduction", "Working_Days"
    ]

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.append(headers)
    header_fill = PatternFill(start_color="1b6656", end_color="1b6656", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = 24

    # Pre-fill with existing employees' default data
    all_emps = get_all_employees()
    for emp in all_emps:
        ws.append([
            emp.get("employee_id", ""),
            emp.get("basic_salary", 0),
            emp.get("medical_allowance", 0),
            emp.get("dearness_allowance", 0),
            0, # Arrears
            emp.get("transport_allowance", 0),
            emp.get("cola_allowance", 0),
            emp.get("utility_allowance", 0),
            0, # Washing allowance
            emp.get("previous_month_allowance", 0),
            emp.get("bonus_allowance", 0),
            emp.get("other_allowance", 0),
            0, # Paid_leave_amount
            0, # Deduction_misc
            emp.get("overtime", 0),
            0, # Damage_medical
            emp.get("income_tax", 0),
            0,   # Unpaid_Leaves
            emp.get("eobi_deduction", 0),
            emp.get("other_deduction", 0),
            26,  # Working_Days
        ])
        for cell in ws[ws.max_row]:
            cell.border = thin

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output, as_attachment=True,
        download_name="DACI_Salary_Upload_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/generate/excel-upload", methods=["GET", "POST"])
@login_required
@hr_required
def generate_from_excel():
    """Upload an Excel file to bulk-generate salary slips."""
    employees = get_all_employees()
    now = datetime.now()

    if request.method == "POST":
        if "excel_file" not in request.files:
            flash("Koi file select nahi ki gayi.", "warning")
            return redirect(request.url)

        file = request.files["excel_file"]
        if not file or file.filename == "":
            flash("Koi file select nahi ki gayi.", "warning")
            return redirect(request.url)

        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            flash("Sirf Excel (.xlsx / .xls) files allowed hain.", "danger")
            return redirect(request.url)

        try:
            import tempfile
            from openpyxl import load_workbook

            # Save to temp file and read
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name

            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active

            # Parse header row → normalise keys
            headers = [
                str(cell.value).strip().lower().replace(" ", "_") if cell.value else ""
                for cell in ws[1]
            ]

            # Employee lookup maps
            emp_by_id   = {str(e["employee_id"]).strip(): e for e in employees}
            emp_by_name = {e["name"].strip().lower(): e  for e in employees}

            # Month/Year from form (override all rows)
            month_override = request.form.get("month", type=int)
            year_override  = request.form.get("year",  type=int)

            success_count = 0
            error_list    = []

            def _f(d, key, fallback=0.0):
                try: return float(d.get(key) or fallback)
                except: return float(fallback)

            def _i(d, key, fallback=0):
                try: return int(d.get(key) or fallback)
                except: return int(fallback)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(v for v in row if v is not None):
                    continue  # skip blank rows

                rd = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

                # Resolve employee
                emp = None
                eid = rd.get("employee_id")
                if eid:
                    emp = emp_by_id.get(str(eid).strip())
                if not emp:
                    nm = rd.get("name") or rd.get("employee_name", "")
                    if nm:
                        emp = emp_by_name.get(str(nm).strip().lower())

                if not emp:
                    error_list.append(f"Row {row_idx}: Employee '{eid}' not found — skipped.")
                    continue

                month = month_override or _i(rd, "month", now.month)
                year  = year_override  or _i(rd, "year",  now.year)

                basic      = _f(rd, "basic_salary",             emp.get("basic_salary", 0))
                medical    = _f(rd, "medical_allowance",        emp.get("medical_allowance", 0))
                house      = _f(rd, "house_allowance",          emp.get("house_allowance", 0))
                transport  = _f(rd, "transport_allowance",      emp.get("transport_allowance", 0))
                dearness   = _f(rd, "dearness_allowance",       emp.get("dearness_allowance", 0))
                cola       = _f(rd, "cola_allowance",           emp.get("cola_allowance", 0))
                utility    = _f(rd, "utility_allowance",        emp.get("utility_allowance", 0))
                washing    = _f(rd, "washing_allowance",        0)
                prev_month = _f(rd, "previous_month_allowance", emp.get("previous_month_allowance", 0))
                bonus      = _f(rd, "bonus_allowance",          emp.get("bonus_allowance", 0))
                other_all  = _f(rd, "other_allowance",          emp.get("other_allowance", 0))
                arrears    = _f(rd, "arrears",                  0)

                gross = basic + medical + house + transport + dearness + cola + utility + washing + prev_month + bonus + other_all + arrears

                leave_enc  = _f(rd, "paid_leave_amount",        0)
                overtime   = _f(rd, "overtime",                 emp.get("overtime", 0))
                ded_misc   = _f(rd, "deduction_misc",           0)
                dmg_med    = _f(rd, "damage_medical",           0)

                taxable = gross + leave_enc + overtime - ded_misc - dmg_med

                tax       = _f(rd, "income_tax",      emp.get("income_tax", 0))
                eobi      = _f(rd, "eobi_deduction",  emp.get("eobi_deduction", 0))
                unpaid    = _f(rd, "unpaid_leaves",   0)
                other_ded = _f(rd, "other_deduction", emp.get("other_deduction", 0))
                total_ded = tax + eobi + unpaid + other_ded
                net       = taxable - total_ded
                w_days    = _i(rd, "working_days", 26)

                slip_data = {
                    "employee_id":              emp["id"],
                    "month":                    month,
                    "year":                     year,
                    "basic_salary":             basic,
                    "medical_allowance":        medical,
                    "dearness_allowance":       dearness,
                    "house_allowance":          house,
                    "transport_allowance":      transport,
                    "cola_allowance":           cola,
                    "utility_allowance":        utility,
                    "washing_allowance":        washing,
                    "previous_month_allowance": prev_month,
                    "bonus_allowance":          bonus,
                    "other_allowance":          other_all,
                    "arrears":                  arrears,
                    "gross_salary":             gross,
                    "paid_leave_amount":        leave_enc,
                    "overtime":                 overtime,
                    "deduction_misc":           ded_misc,
                    "damage_medical":           dmg_med,
                    "taxable_salary":           taxable,
                    "income_tax":               tax,
                    "eobi_deduction":           eobi,
                    "unpaid_leaves":            unpaid,
                    "other_deduction":          other_ded,
                    "total_deductions":         total_ded,
                    "net_salary":               net,
                    "working_days":             w_days,
                    "generated_by":             current_user.email,
                    "generated_at":             datetime.now().isoformat(),
                }

                try:
                    pdf_path = generate_and_upload_slip(slip_data, emp)
                    slip_data["pdf_path"] = pdf_path
                    save_salary_slip(slip_data)
                    success_count += 1
                except Exception as ex:
                    error_list.append(f"Row {row_idx} ({emp['name']}): {str(ex)}")

            # Cleanup temp
            try: os.unlink(tmp_path)
            except: pass

            if success_count:
                log_activity(current_user.email, "Excel Upload",
                             f"Generated {success_count} slips via Excel for {MONTHS[month_override or now.month]} {year_override or now.year}")
                flash(f"✅ {success_count} salary slip(s) successfully generated from Excel!", "success")
            if error_list:
                flash(f"⚠️ {len(error_list)} row(s) skipped: " + " | ".join(error_list[:5]), "warning")

            return redirect(url_for("view_slips"))

        except Exception as e:
            flash(f"❌ Excel parsing error: {str(e)}", "danger")
            return redirect(request.url)

    return render_template(
        "excel_upload.html",
        employees=employees,
        months=MONTHS,
        current_month=now.month,
        current_year=now.year,
        years=range(2023, 2041)
    )


@app.route("/logs")
@login_required
@admin_required
def view_logs():
    res = supabase.table("activity_logs").select("*").order("timestamp", desc=True).limit(200).execute()
    logs = res.data or []
    return render_template("activity_logs.html", logs=logs)


# ── User Management (Admin Only) ──────────────────────────────────

@app.route("/users")
@login_required
@admin_required
def manage_users():
    res = supabase.table("users").select("*").order("id").execute()
    users = res.data or []
    return render_template("manage_users.html", users=users)


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@admin_required
def add_user():
    if request.method == "POST":
        email    = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        role     = request.form.get("role", "hr")

        from utils.db import get_user_by_email, create_user
        existing = get_user_by_email(email)
        if existing:
            flash("Yeh email pehle se exist karta hai!", "danger")
        else:
            hashed = bcrypt.generate_password_hash(password).decode("utf-8")
            create_user(email, hashed, role)
            log_activity(current_user.email, "Add User", f"Created new user: {email} with role: {role}")
            flash(f"✅ User {email} ({role}) successfully add ho gaya!", "success")
            return redirect(url_for("manage_users"))

    return render_template("add_user.html")


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):
    if user_id == current_user.id:
        flash("Aap apna khud ka account delete nahi kar sakte!", "danger")
        return redirect(url_for("manage_users"))
    from utils.db import supabase
    # Get user email before deleting
    user_res = supabase.table("users").select("email").eq("id", user_id).single().execute()
    target_email = user_res.data['email'] if user_res.data else "Unknown"
    
    supabase.table("users").delete().eq("id", user_id).execute()
    log_activity(current_user.email, "Delete User", f"Deleted user account: {target_email}")
    flash("User delete ho gaya.", "info")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:user_id>/change-role", methods=["POST"])
@login_required
@admin_required
def change_role(user_id):
    if user_id == current_user.id:
        flash("Aap apna khud ka role change nahi kar sakte!", "danger")
        return redirect(url_for("manage_users"))
    new_role = request.form.get("role")
    from utils.db import supabase
    # Get user email before updating
    user_res = supabase.table("users").select("email").eq("id", user_id).single().execute()
    target_email = user_res.data['email'] if user_res.data else "Unknown"
    
    supabase.table("users").update({"role": new_role}).eq("id", user_id).execute()
    log_activity(current_user.email, "Change Role", f"Changed role of {target_email} to {new_role}")
    flash(f"Role successfully update ho gaya!", "success")
    return redirect(url_for("manage_users"))


@app.errorhandler(404)
def not_found_error(error):
    return render_template("errors/404.html"), 404

@app.errorhandler(403)
def forbidden_error(error):
    return render_template("errors/403.html"), 403

@app.errorhandler(500)
def internal_error(error):
    log_activity("SYSTEM", "500 Error", str(error))
    return render_template("errors/500.html"), 500

if __name__ == "__main__":
    app.run(debug=True, port=5000)